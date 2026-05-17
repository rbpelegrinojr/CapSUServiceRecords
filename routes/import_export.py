import io
import os
import re
from datetime import datetime

import openpyxl
import pandas as pd
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from werkzeug.utils import secure_filename
from database import db
from models.employee import Employee
from models.service_record import ServiceRecord

import_export_bp = Blueprint('import_export', __name__)

ALLOWED_EXTENSIONS = {'xlsx'}

# Row-scan limits used when parsing the official CapSU Service Records form
_MAX_EMP_INFO_SEARCH_ROWS = 50   # rows to search for the NAME / BIRTH labels
_MAX_HEADER_SEARCH_ROWS = 35     # rows after NAME row to search for the FROM/TO sub-header
_MAX_EMP_INFO_SCAN_COLUMN = 12
_HEADER_LOOKBACK_ROWS = 2
_MAX_HEADER_SCAN_COLUMN = 20
_MAX_DISPLAYED_ERRORS = 10


def _allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _parse_official_format(file_bytes, filename):
    """
    Parse a single official CapSU Service Records .xlsx file (one file per employee).

    Expected layout:
      - A row with "NAME" in col A, followed by Surname (col B), Given Name (col C),
        Middle Name (col D).
      - A row with "BIRTH" in col A, followed by birth date (col B) and birth place (col D).
      - A sub-header row containing "FROM" somewhere; data rows follow immediately after.

    Returns (emp_data dict, records list, error_message string).
    On failure emp_data and records are None and error_message is set.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        return None, None, f'{filename}: Could not open file – {exc}'

    ws = wb.active

    def cv(row, col):
        """Return cell value as a stripped string, formatting dates as MM/DD/YYYY."""
        v = ws.cell(row=row, column=col).value
        if v is None:
            return ''
        if hasattr(v, 'strftime'):
            return v.strftime('%m/%d/%Y')
        return str(v).strip()

    # ── Locate the NAME and BIRTH rows ──────────────────────────────────────
    name_row = birth_row = None
    for r in range(1, min(ws.max_row + 1, _MAX_EMP_INFO_SEARCH_ROWS)):
        a = cv(r, 1).upper()
        if a == 'NAME':
            name_row = r
        elif a == 'BIRTH':
            birth_row = r

    if name_row is None:
        return None, None, (
            f'{filename}: Could not find a "NAME" label in column A. '
            'Make sure the file uses the standard CapSU Service Records form.'
        )

    # ── Read employee information ────────────────────────────────────────────
    # Handle both strict B/C/D layout and merged-cell official template layout.
    max_emp_info_col = min(ws.max_column, _MAX_EMP_INFO_SCAN_COLUMN)
    def safe_upper(value):
        return (value or '').upper()

    name_values = []
    for c in range(2, max_emp_info_col + 1):
        value = cv(name_row, c)
        if value:
            name_values.append(value)
    surname = safe_upper(name_values[0] if len(name_values) >= 1 else cv(name_row, 2))
    given_name = safe_upper(name_values[1] if len(name_values) >= 2 else cv(name_row, 3))
    middle_name = safe_upper(name_values[2] if len(name_values) >= 3 else cv(name_row, 4)) or None

    birth_values = []
    if birth_row:
        for c in range(2, max_emp_info_col + 1):
            value = cv(birth_row, c)
            if value:
                birth_values.append(value)
    birth_date = birth_values[0] if len(birth_values) >= 1 else (cv(birth_row, 2) if birth_row else '')
    birth_place = birth_values[1] if len(birth_values) >= 2 else (cv(birth_row, 4) if birth_row else '')

    if not surname or not given_name:
        return None, None, (
            f'{filename}: Could not read Surname or Given Name from the NAME row (row {name_row}). '
            'Expected: col B = Surname, col C = Given Name.'
        )

    # ── Locate the sub-header row that contains "FROM" ──────────────────────
    from_col = to_col = desig_col = status_col = None
    salary_col = station_col = branch_col = lv_col = None
    sep_date_col = sep_cause_col = None
    data_start_row = None

    search_from = name_row + 1
    search_to = min(ws.max_row, name_row + _MAX_HEADER_SEARCH_ROWS)
    for r in range(search_from, search_to + 1):
        row_upper = [cv(r, c).upper().replace('\n', ' ') for c in range(1, 13)]
        if 'FROM' not in row_upper:
            continue

        # Some official files split the header across multiple rows.
        header_start_row = max(1, r - _HEADER_LOOKBACK_ROWS)
        header_end_row = min(ws.max_row, r + 1)
        header_rows = [hr for hr in range(header_start_row, header_end_row + 1)]
        max_scan_col = min(ws.max_column, _MAX_HEADER_SCAN_COLUMN)
        for ci in range(1, max_scan_col + 1):
            parts = []
            for hr in header_rows:
                value = cv(hr, ci)
                if value:
                    parts.append(value.upper().replace('\n', ' '))
            labels = ' '.join(parts)
            if not labels:
                continue
            if 'FROM' in labels and from_col is None:
                from_col = ci
            # Use strict standalone matching for TO to avoid false positives (e.g., in STATION).
            elif re.search(r'\bTO\b', labels) and to_col is None:
                to_col = ci
            elif 'DESIGNATION' in labels and desig_col is None:
                desig_col = ci
            elif 'STATUS' in labels and status_col is None:
                status_col = ci
            elif 'SALARY' in labels and salary_col is None:
                salary_col = ci
            elif ('STATION' in labels or 'PLACE' in labels) and station_col is None:
                station_col = ci
            elif 'BRANCH' in labels and branch_col is None:
                branch_col = ci
            elif ('PAY' in labels or 'W/O' in labels) and lv_col is None:
                lv_col = ci
            elif 'SEPARATION' in labels and 'DATE' in labels and sep_date_col is None:
                sep_date_col = ci
            elif 'CAUSE' in labels and sep_cause_col is None:
                sep_cause_col = ci
        data_start_row = r + 1
        break

    # Fallback to fixed column positions if header detection failed
    if data_start_row is None:
        from_col, to_col = 1, 2
        desig_col, status_col = 3, 4
        salary_col = 5
        station_col, branch_col = 6, 7
        lv_col = 8
        sep_date_col, sep_cause_col = 9, 10
        data_start_row = (birth_row + 10) if birth_row else 24

    # ── Read service record rows ─────────────────────────────────────────────
    records = []
    last_used_col = max(
        c for c in (from_col, to_col, desig_col, status_col, salary_col,
                    station_col, branch_col, lv_col, sep_date_col, sep_cause_col)
        if c is not None
    )

    for r in range(data_start_row, ws.max_row + 1):
        row_vals = [cv(r, c) for c in range(from_col, last_used_col + 1)]
        if not any(row_vals):
            continue

        from_val = cv(r, from_col)

        # Skip rows that look like repeated headers or stray labels
        if from_val.upper() in ('FROM', 'SERVICE', '(INCLUSIVE DATES)', ''):
            continue

        salary_raw = cv(r, salary_col) if salary_col else ''
        monthly_salary = None
        try:
            m = re.search(r'\d+\.?\d*', salary_raw.replace(',', ''))
            if m:
                monthly_salary = float(m.group().replace(',', ''))
        except (ValueError, TypeError):
            pass

        records.append({
            'date_from': from_val,
            'date_to': cv(r, to_col) if to_col else '',
            'designation': cv(r, desig_col) if desig_col else '',
            'status': cv(r, status_col) if status_col else '',
            'monthly_salary': monthly_salary,
            'annual_salary': None,
            'station_place_of': cv(r, station_col) if station_col else '',
            'branch': cv(r, branch_col) if branch_col else '',
            'lv_abs_wo_pay': cv(r, lv_col) if lv_col else '',
            'separation_date': cv(r, sep_date_col) if sep_date_col else '',
            'separation_cause': cv(r, sep_cause_col) if sep_cause_col else '',
        })

    emp_data = {
        'surname': surname,
        'given_name': given_name,
        'middle_name': middle_name,
        'birth_date': birth_date,
        'birth_place': birth_place,
    }
    return emp_data, records, None


@import_export_bp.route('/import', methods=['GET', 'POST'])
def import_records():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part in the request.', 'danger')
            return redirect(url_for('import_export.import_records'))
        f = request.files['file']
        if f.filename == '':
            flash('No file selected.', 'danger')
            return redirect(url_for('import_export.import_records'))
        if not _allowed_file(f.filename):
            flash('Only .xlsx files are allowed.', 'danger')
            return redirect(url_for('import_export.import_records'))

        try:
            df = pd.read_excel(io.BytesIO(f.read()), dtype=str)
            df.columns = [c.strip() for c in df.columns]
            df = df.where(pd.notnull(df), None)

            required_cols = {'Surname', 'Given Name'}
            missing = required_cols - set(df.columns)
            if missing:
                flash(f'Missing required columns: {", ".join(missing)}', 'danger')
                return redirect(url_for('import_export.import_records'))

            created_employees = 0
            created_records = 0
            errors = []

            for idx, row in df.iterrows():
                try:
                    surname = (row.get('Surname') or '').strip().upper()
                    given_name = (row.get('Given Name') or '').strip().upper()
                    if not surname or not given_name:
                        errors.append(f'Row {idx + 2}: Missing Surname or Given Name.')
                        continue

                    emp = Employee.query.filter(
                        db.func.upper(Employee.surname) == surname,
                        db.func.upper(Employee.given_name) == given_name,
                    ).first()

                    if not emp:
                        emp = Employee(
                            surname=surname,
                            given_name=given_name,
                            middle_name=(row.get('Middle Name') or '').strip().upper() or None,
                            birth_date=row.get('Birth Date'),
                            birth_place=row.get('Birth Place'),
                        )
                        db.session.add(emp)
                        db.session.flush()
                        created_employees += 1

                    def _safe_float(val):
                        try:
                            return float(val) if val not in (None, '') else None
                        except (ValueError, TypeError):
                            return None

                    max_sort = db.session.query(
                        db.func.max(ServiceRecord.sort_order)
                    ).filter_by(employee_id=emp.id).scalar() or 0

                    rec = ServiceRecord(
                        employee_id=emp.id,
                        date_from=row.get('Date From'),
                        date_to=row.get('Date To'),
                        designation=row.get('Designation'),
                        status=row.get('Status'),
                        monthly_salary=_safe_float(row.get('Monthly Salary')),
                        annual_salary=_safe_float(row.get('Annual Salary')),
                        station_place_of=row.get('Station/Place of'),
                        branch=row.get('Branch'),
                        lv_abs_wo_pay=row.get('LV ABS WO Pay'),
                        separation_date=row.get('Separation Date'),
                        separation_cause=row.get('Separation Cause'),
                        sort_order=max_sort + 1,
                    )
                    db.session.add(rec)
                    created_records += 1

                except Exception as e:
                    errors.append(f'Row {idx + 2}: {str(e)}')

            db.session.commit()

            msg = f'Import complete: {created_employees} employee(s) created, {created_records} record(s) imported.'
            flash(msg, 'success')
            if errors:
                for err in errors[:_MAX_DISPLAYED_ERRORS]:
                    flash(err, 'warning')
                if len(errors) > _MAX_DISPLAYED_ERRORS:
                    flash(f'... and {len(errors) - _MAX_DISPLAYED_ERRORS} more errors.', 'warning')

        except Exception as e:
            db.session.rollback()
            flash(f'Import failed: {str(e)}', 'danger')

        return redirect(url_for('import_export.import_records'))

    return render_template('import_export/import.html')


@import_export_bp.route('/import-official', methods=['POST'])
def import_official_records():
    """Import one or more per-employee official CapSU Service Records .xlsx files."""
    files = request.files.getlist('files')
    if not files or all(f.filename == '' for f in files):
        flash('No files selected.', 'danger')
        return redirect(url_for('import_export.import_records'))

    total_employees = 0
    total_records = 0
    all_errors = []

    for f in files:
        if f.filename == '':
            continue
        if not _allowed_file(f.filename):
            all_errors.append(f'{f.filename}: Only .xlsx files are allowed.')
            continue

        emp_data, records, err = _parse_official_format(f.read(), f.filename)
        if err:
            all_errors.append(err)
            continue

        try:
            surname = emp_data['surname']
            given_name = emp_data['given_name']

            emp = Employee.query.filter(
                db.func.upper(Employee.surname) == surname,
                db.func.upper(Employee.given_name) == given_name,
            ).first()

            if not emp:
                emp = Employee(
                    surname=surname,
                    given_name=given_name,
                    middle_name=emp_data.get('middle_name'),
                    birth_date=emp_data.get('birth_date') or None,
                    birth_place=emp_data.get('birth_place') or None,
                )
                db.session.add(emp)
                db.session.flush()
                total_employees += 1

            max_sort = db.session.query(
                db.func.max(ServiceRecord.sort_order)
            ).filter_by(employee_id=emp.id).scalar() or 0

            for i, rec_data in enumerate(records, start=1):
                rec = ServiceRecord(
                    employee_id=emp.id,
                    date_from=rec_data['date_from'] or None,
                    date_to=rec_data['date_to'] or None,
                    designation=rec_data['designation'] or None,
                    status=rec_data['status'] or None,
                    monthly_salary=rec_data['monthly_salary'],
                    annual_salary=rec_data['annual_salary'],
                    station_place_of=rec_data['station_place_of'] or None,
                    branch=rec_data['branch'] or None,
                    lv_abs_wo_pay=rec_data['lv_abs_wo_pay'] or None,
                    separation_date=rec_data['separation_date'] or None,
                    separation_cause=rec_data['separation_cause'] or None,
                    sort_order=max_sort + i,
                )
                db.session.add(rec)
                total_records += 1

            db.session.commit()

        except Exception as exc:
            db.session.rollback()
            all_errors.append(f'{f.filename}: {exc}')

    if total_employees or total_records:
        flash(
            f'Import complete: {total_employees} employee(s) created, '
            f'{total_records} record(s) imported.',
            'success',
        )
    for err in all_errors[:_MAX_DISPLAYED_ERRORS]:
        flash(err, 'warning')
    if len(all_errors) > _MAX_DISPLAYED_ERRORS:
        flash(f'… and {len(all_errors) - _MAX_DISPLAYED_ERRORS} more errors.', 'warning')
    if not (total_employees or total_records) and not all_errors:
        flash('No records were found in the uploaded file(s).', 'warning')

    return redirect(url_for('import_export.import_records'))


@import_export_bp.route('/combine-files', methods=['POST'])
def combine_files():
    """Combine multiple official-format files into one flat system-format file."""
    files = request.files.getlist('files')
    if not files or all(f.filename == '' for f in files):
        flash('No files selected.', 'danger')
        return redirect(url_for('import_export.import_records'))

    rows = []
    all_errors = []

    for f in files:
        if f.filename == '':
            continue
        if not _allowed_file(f.filename):
            all_errors.append(f'{f.filename}: Only .xlsx files are allowed.')
            continue

        emp_data, records, err = _parse_official_format(f.read(), f.filename)
        if err:
            all_errors.append(err)
            continue

        for rec in records:
            rows.append({
                'Surname': emp_data['surname'],
                'Given Name': emp_data['given_name'],
                'Middle Name': emp_data.get('middle_name') or '',
                'Birth Date': emp_data.get('birth_date') or '',
                'Birth Place': emp_data.get('birth_place') or '',
                'Date From': rec['date_from'] or '',
                'Date To': rec['date_to'] or '',
                'Designation': rec['designation'] or '',
                'Status': rec['status'] or '',
                'Monthly Salary': rec['monthly_salary'] if rec['monthly_salary'] is not None else '',
                'Annual Salary': rec['annual_salary'] if rec['annual_salary'] is not None else '',
                'Station/Place of': rec['station_place_of'] or '',
                'Branch': rec['branch'] or '',
                'LV ABS WO Pay': rec['lv_abs_wo_pay'] or '',
                'Separation Date': rec['separation_date'] or '',
                'Separation Cause': rec['separation_cause'] or '',
            })

    for err in all_errors[:_MAX_DISPLAYED_ERRORS]:
        flash(err, 'warning')
    if len(all_errors) > _MAX_DISPLAYED_ERRORS:
        flash(f'… and {len(all_errors) - _MAX_DISPLAYED_ERRORS} more errors.', 'warning')

    if not rows:
        flash('No records could be read from the uploaded file(s).', 'danger')
        return redirect(url_for('import_export.import_records'))

    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Service Records')
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"combined_service_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )


@import_export_bp.route('/export/<int:employee_id>')
def export_employee(employee_id):
    emp = Employee.query.get_or_404(employee_id)
    records = ServiceRecord.query.filter_by(employee_id=employee_id).order_by(
        ServiceRecord.sort_order
    ).all()

    rows = []
    for rec in records:
        rows.append({
            'Surname': emp.surname,
            'Given Name': emp.given_name,
            'Middle Name': emp.middle_name or '',
            'Birth Date': emp.birth_date or '',
            'Birth Place': emp.birth_place or '',
            'Date From': rec.date_from or '',
            'Date To': rec.date_to or '',
            'Designation': rec.designation or '',
            'Status': rec.status or '',
            'Monthly Salary': rec.monthly_salary or '',
            'Annual Salary': rec.annual_salary or '',
            'Station/Place of': rec.station_place_of or '',
            'Branch': rec.branch or '',
            'LV ABS WO Pay': rec.lv_abs_wo_pay or '',
            'Separation Date': rec.separation_date or '',
            'Separation Cause': rec.separation_cause or '',
        })

    if not rows:
        rows = [{'Surname': emp.surname, 'Given Name': emp.given_name}]

    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Service Records')
    output.seek(0)

    filename = f'service_records_{emp.surname}_{emp.given_name}.xlsx'.replace(' ', '_')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@import_export_bp.route('/download-sample-template')
def download_sample_template():
    sample_path = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),
        'sample_data',
        'sample_service_records.xlsx',
    )
    if not os.path.exists(sample_path):
        flash('Sample template not found.', 'warning')
        return redirect(url_for('import_export.import_records'))
    return send_file(
        sample_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='sample_service_records.xlsx',
    )
